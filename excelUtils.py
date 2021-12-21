from openpyxl.styles import Alignment, PatternFill

fill = [{
        0: PatternFill("solid", fgColor="b5cfb4"),
        1: PatternFill("solid", fgColor="90b78f"),
        2: PatternFill("solid", fgColor="77a776")
    }, {
        0: PatternFill("solid", fgColor="f6cecb"),
        1: PatternFill("solid", fgColor="f0ada8"),
        2: PatternFill("solid", fgColor="ea8c85")
    }
    ]


# 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
def get_num_colnum_dict():
    '''
:return: 返回字典：{1:'A', 2:'B', ...... , 52:'AZ'}
'''
    num_str_dict = {}
    A_Z = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
    AA_AZ = ['A' + chr(a) for a in range(ord('A'), ord('Z') + 1)]
    A_AZ = A_Z + AA_AZ
    for i in A_AZ:
        num_str_dict[A_AZ.index(i) + 1] = i
    return num_str_dict


# 自适应列宽
def style_excel(sheet):
    '''
    :param sheet_name:  excel中的sheet名
    :return:
    '''
    # 选择对应的sheet
    # 获取最大行数与最大列数
    max_column = sheet.max_column
    max_row = sheet.max_row

    # 将每一列，单元格列宽最大的列宽值存到字典里，key:列的序号从1开始(与字典num_str_dic中的key对应)；value:列宽的值
    max_column_dict = {}

    # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
    num_str_dict = get_num_colnum_dict()

    # 遍历全部列
    for i in range(1, max_column + 1):
        # 遍历每一列的全部行
        for j in range(1, max_row + 1):
            #   为每个单元格添加居中样式
            align = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=j, column=i).alignment = align
            column = 0
            # 获取j行i列的值
            sheet_value = sheet.cell(row=j, column=i).value
            # 通过列表生成式生成字符列表，将当前获取到的单元格的str值的每一个字符放在一个列表中（列表中一个元素是一个字符）
            sheet_value_list = [k for k in str(sheet_value)]
            # 遍历当前单元格的字符列表
            for v in sheet_value_list:
                # 判定长度，一个数字或一个字母，单元格列宽+=1.1，其它+=2.2（长度可根据需要自行修改，经测试一个字母的列宽长度大概为1）
                if ord(v) <= 256:
                    column += 1.1
                else:
                    column += 2.2
            # 当前单元格列宽与字典中的对比，大于字典中的列宽值则将字典更新。如果字典没有这个key，抛出异常并将值添加到字典中
            try:
                if column > max_column_dict[i]:
                    max_column_dict[i] = column
            except Exception as e:
                max_column_dict[i] = column
    # 此时max_column_dict字典中已存有当前sheet的所有列的最大列宽值，直接遍历字典修改列宽
    for key, value in max_column_dict.items():
        sheet.column_dimensions[num_str_dict[key]].width = value + 1


# isLost = 0 代表win，否则 lost
def fillColorByValue(Max, num, sheet, r, c, isLost):
    if num == 0:
        return
    print(Max, num)
    scope = Max / 3
    if 0 < num < scope:
        sheet.cell(r, c).fill = fill[isLost][0]
    elif scope <= num < 2 * scope:
        sheet.cell(r, c).fill = fill[isLost][1]
    else:
        sheet.cell(r, c).fill = fill[isLost][2]
