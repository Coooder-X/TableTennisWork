# 返回本 point 发球的 player 对象
import os
from openpyxl.styles import Alignment, PatternFill, Font
import excelUtils


def getDoublesServeSide(point, playerList):
    rallyList = point['list']
    serveStr = rallyList[0]['HitPlayer']
    teamID, playerID = int(serveStr[0]), int(serveStr[1])
    # print(playerList[teamID][playerID]['name'])
    return playerList[teamID][playerID]


def getPlayerName(pair, playerList):
    teamID, playerID = int(pair[0]), int(pair[1])
    return playerList[teamID][playerID]['name']


def judgeIsDouble(playerList):
    if len(playerList[0]) == 2:
        return True
    return False


def getHitOrders(playerList):
    orderList = []
    if judgeIsDouble(playerList):
        playerIDList = ['00', '01', '10', '11']
        for i in range(2):
            for j in range(2, 4):
                orderList.append([playerIDList[i], playerIDList[j]])
                orderList.append([playerIDList[j], playerIDList[i]])
    else:
        orderList = [['00', '10'], ['10', '00']]
    # print(orderList.index(['10', '00']))
    return orderList


# 返回一个长度为2的列表，第一个元素表示作为统计的对象的队伍ID，第二个是对方的队伍ID
# 如果一支中国一支国外，中国为统计对象，若两支中国或两支国外，选第一个
# 注意返回的是字符串代表队伍类型
def getTargetTeamID(playerList):
    cnt = 0
    res = 0
    for idx in range(2):
        team = playerList[idx]
        player = team[0]
        print(player)
        if player['country'] == '中国':
            cnt += 1
            res = idx
    if cnt == 1:
        return [str(res), '0' if res == 1 else '1']
    else:  # cnt == 0 or cnt == 2:
        return ['0', '1']


# hitPair是形如 ['00', '10'] 的运动员编号对，本函数返回的是中国运动员在本列表中的下标
def getTargetPlayer(hitPair, targetTeamID):
    for i in range(2):
        if hitPair[i][0] == targetTeamID:
            return i


def getOppositePlayer(hitPair, targetTeamID):
    for i in range(2):
        if hitPair[i][0] != targetTeamID:
            return i


def updateScoreCase(scoreCase, index, beginType, num, res):
    if num not in scoreCase[index][beginType].keys():
        scoreCase[index][beginType][num] = {'win': 0, 'lost': 0}
    scoreCase[index][beginType][num][res] += 1


def getTeamMate(playerID):
    res = playerID[0]
    res += '0' if playerID[1] == '1' else '1'
    return res


def createDoublesExcel(excelScoreCase, calScore):
    pass


#   carry 双打换发球的情况，第一个if是开局初始化顺序，第二个是局间、决胜局第5分换次序的情况
def updateServeRecOrder(serve_rec_order, serve, rec=''):
    if serve_rec_order['00'] == '':  # 第一局第一分的初始化
        serve_rec_order[serve] = rec
        serve_rec_order[rec] = getTeamMate(serve)
        serve_rec_order[getTeamMate(serve)] = getTeamMate(rec)
        serve_rec_order[getTeamMate(rec)] = serve
        return
    if rec == '':
        rec = serve_rec_order[serve]
        serve_rec_order[serve], serve_rec_order[getTeamMate(serve)] = serve_rec_order[getTeamMate(serve)], \
                                                                      serve_rec_order[serve]
        serve_rec_order[rec], serve_rec_order[getTeamMate(rec)] = serve_rec_order[getTeamMate(rec)], serve_rec_order[
            rec]
        # serve_rec_order[serve] = getTeamMate(serve_rec_order[serve])
        # serve_rec_order[rec] = serve
        # serve_rec_order[getTeamMate(serve)] = rec
        # serve_rec_order[getTeamMate(rec)] = getTeamMate(serve)


def updateDoublesData(scoreCase, rallyList, hitOrders, TargetTeamPair, serve_rec_order, isOpposite, cnt, cnt2):
    for index, hitPair in enumerate(hitOrders):
        rallyNum = len(rallyList)

        A1 = hitPair[getTargetPlayer(hitPair, TargetTeamPair[0])]  # 当前pair中中国运动员
        B1 = hitPair[getOppositePlayer(hitPair, TargetTeamPair[0])]  # 当前pair对方运动员

        if isOpposite:
            A1, B1 = B1, A1

        if rallyNum == 1:  # 发球失误的情况特判一下
            pair = [rallyList[0]['HitPlayer'], serve_rec_order[rallyList[0]['HitPlayer']]]
            idx = hitOrders.index(pair)
            print('发球失误', cnt, ':', cnt2)
            if pair[0][0] == A1[0]:
                print('本方发球失误', pair)
                cnt2 += 1
                updateScoreCase(scoreCase, idx, 0, 0, 'lost')
            elif pair[0][0] == B1[0]:
                print('对方发球失误', pair)
                cnt += 1
                updateScoreCase(scoreCase, idx, 0, 1, 'win')
            else:
                print(A1, B1, '|', pair)
            break

        servePair = [rallyList[0]['HitPlayer'], rallyList[1]['HitPlayer']]

        lastHitPair = [rallyList[-2]['HitPlayer'], rallyList[-1]['HitPlayer']]
        # 满足A1->B1的模式，计算A1给B1赢的情况，分2种：A1发球和A2发球
        if servePair[0][0] == A1[0] and hitPair[0][0] == A1[0]:  # 必须要满足A发球并且hitPair也是A在前，否则会出现拍数归类错误
            if hitPair == lastHitPair and A1 == hitPair[0]:
                # if servePair[0][0] == A1[0]:
                if servePair[0] == A1:
                    cnt += 1
                    updateScoreCase(scoreCase, index, 0, rallyNum - 2, 'win')
                    break
                else:
                    cnt += 1
                    updateScoreCase(scoreCase, index, 1, rallyNum - 2, 'win')
                    break
            # 满足A1->B1的模式，计算A1给B1输的情况，分2种：A1发球和A2发球，最后一rally是B2->A1
            elif lastHitPair[0] != B1 and lastHitPair[1] == A1:
                # if servePair[0][0] == A1[0]:
                if servePair[0] == A1:
                    cnt2 += 1
                    updateScoreCase(scoreCase, index, 0, rallyNum - 1, 'lost')
                    break
                else:
                    cnt2 += 1
                    updateScoreCase(scoreCase, index, 1, rallyNum - 1, 'lost')
                    break

        elif servePair[0][0] == B1[0] and hitPair[0][0] == B1[0]:
            # 满足B1->A1的模式，计算B1给A1赢（A1输）的情况，分2种：B1发球和B2发球
            if lastHitPair == hitPair and hitPair[0] == B1:
                # if servePair[0][0] == B1[0]:
                if servePair[0] == B1:
                    cnt2 += 1
                    updateScoreCase(scoreCase, index, 0, rallyNum - 1, 'lost')
                    break
                else:
                    cnt2 += 1
                    updateScoreCase(scoreCase, index, 1, rallyNum - 1, 'lost')
                    break
            # 满足B1->A1的模式，计算B1给A1输（A1赢）的情况，分2种：B1发球和B2发球，最后一rally是A1->B2
            elif lastHitPair[0] == A1 and lastHitPair[1] != B1:
                # if servePair[0][0] == B1[0]:
                if servePair[0] == B1:
                    cnt += 1
                    updateScoreCase(scoreCase, index, 0, rallyNum - 2, 'win')
                    break
                else:
                    cnt += 1
                    updateScoreCase(scoreCase, index, 1, rallyNum - 2, 'win')
                    break


def getExcelScoreCase(excelScoreCase, scoreCase, calScore, hitOrders, playerList):
    MaxScore = 0
    count = 0
    for i in range(len(scoreCase)):
        # print(hitOrders[i])
        p1 = getPlayerName(hitOrders[i][0], playerList)
        p2 = getPlayerName(hitOrders[i][1], playerList)
        case = p1 + '→' + p2
        case2 = case
        excelScoreCase[case] = {0: {}, 1: {}}
        if len(hitOrders) == 8:  # 区分单、双打情况的输出
            p3 = getPlayerName(getTeamMate(hitOrders[i][0]), playerList)
            p4 = getPlayerName(getTeamMate(hitOrders[i][1]), playerList)
            case2 += '→' + p3 + '→' + p4
        calScore[case2] = {'得分数': 0, '失分数': 0, '得分率': 0, '优劣势': ''}
        win = 0
        lost = 0
        for j in range(2):
            # -------------这段代码用于把某情况的最大拍数以内补齐，即没数据的补0。若还不足3列，则补到3列
            maxShoot = 0
            if len(scoreCase[i][j].keys()) != 0:  # 有可能某个A->B的轮次
                maxShoot = max(scoreCase[i][j].keys())
            tmp = maxShoot
            # print('scoreCase[i][j]', scoreCase[i][j], maxShoot)
            while tmp >= 0:
                if tmp not in scoreCase[i][j].keys():
                    scoreCase[i][j][tmp] = {'win': 0, 'lost': 0}
                    # print('add')
                tmp -= 4
            while len(scoreCase[i][j].keys()) < 3:
                maxShoot += 4
                scoreCase[i][j][maxShoot] = {'win': 0, 'lost': 0}
            # ---------------
            for key in sorted(scoreCase[i][j]):
                if key == 0:
                    shoot = '发球'
                elif key == 1:
                    shoot = '接发球'
                else:
                    shoot = '第' + str(key + 1) + '拍'
                excelScoreCase[case][j][shoot] = scoreCase[i][j][key]
                win += scoreCase[i][j][key]['win']
                lost += scoreCase[i][j][key]['lost']
                MaxScore = max(MaxScore, scoreCase[i][j][key]['win'], scoreCase[i][j][key]['lost'])
                count += scoreCase[i][j][key]['win'] + scoreCase[i][j][key]['lost']
    print(excelScoreCase)
    print(count)
    return MaxScore


def getAdvanceRate(excelScoreCase, scoreCase, calScore, hitOrders, playerList):
    # ----------------- 每局下方不同轮次的得分情况、得分率统计 ------------------
    for (i, beginPair) in enumerate(hitOrders):
        win = 0
        lost = 0
        case = ''
        if len(hitOrders) == 8:  # 区分单、双打情况的输出
            p3 = getTeamMate(beginPair[0])
            p4 = getTeamMate(beginPair[1])
            case = getPlayerName(beginPair[0], playerList) + '→' + getPlayerName(
                beginPair[1], playerList) + '→' + \
                   getPlayerName(p3, playerList) + '→' + getPlayerName(p4, playerList)
            j = hitOrders.index([p3, p4])  # 后两个击球的运动员组合在hitOrders里的下标
            for key in scoreCase[i][0].keys():  # 前两个击球的运动员的得分情况，一定是发球或者接发球的，因此下标是0
                win += scoreCase[i][0][key]['win']
                lost += scoreCase[i][0][key]['lost']
            for key in scoreCase[j][1].keys():  # 后两个击球的运动员的得分情况，一定是第三或第四拍的，因此下标是1
                win += scoreCase[j][1][key]['win']
                lost += scoreCase[j][1][key]['lost']
        else:  # 单打
            case = getPlayerName(beginPair[0], playerList) + '→' + getPlayerName(
                beginPair[1], playerList)
            for key in scoreCase[i][0].keys():
                win += scoreCase[i][0][key]['win']
                lost += scoreCase[i][0][key]['lost']
        calScore[case]['得分数'] = win
        calScore[case]['失分数'] = lost
        if win + lost == 0:
            rate = 0
        else:
            rate = float(win * 100) / float(win + lost)
        calScore[case]['得分率'] = format(rate, '.1f') + '%'
        if rate > 50:
            calScore[case]['优劣势'] = '优势'
        elif rate == 50:
            calScore[case]['优劣势'] = '均势'
        else:
            calScore[case]['优劣势'] = '劣势'
    print(calScore)

    for i in excelScoreCase.keys():
        print(i)
        print('A', excelScoreCase[i][0])
        print('B', excelScoreCase[i][1])
        print('---------------------------------------')


def fillExcel(excelScoreCase, calScore, nowLine, baseColumn, sheet, fileName, MaxScore):
    align = Alignment(horizontal='center', vertical='center')
    for pair in excelScoreCase.keys():
        sheet.cell(nowLine, baseColumn, os.path.split(fileName)[1].split('.')[0] + '   ' + pair).alignment = align
        sheet.cell(nowLine, baseColumn).font = Font(bold=True)
        sheet.merge_cells(start_row=nowLine, start_column=baseColumn, end_row=nowLine, end_column=baseColumn+13)

        doubleLine = excelScoreCase[pair]
        for i in range(2):
            line = doubleLine[i]
            shoots = list(line.keys())
            for j in range(len(shoots)):
                sheet.cell(nowLine + 1 + i * 3, j * 2 + baseColumn, shoots[j]).alignment = align
                sheet.merge_cells(start_row=nowLine + 1 + i * 3, start_column=j * 2 + baseColumn,
                                  end_row=nowLine + 1 + i * 3, end_column=j * 2 + baseColumn + 1)
                sheet.cell(nowLine + 2 + i * 3, j * 2 + baseColumn, '得分').alignment = align
                sheet.cell(nowLine + 2 + i * 3 + 1, j * 2 + baseColumn, line[shoots[j]]['win']).alignment = align
                excelUtils.fillColorByValue(MaxScore, line[shoots[j]]['win'], sheet, nowLine + 2 + i * 3 + 1,
                                            j * 2 + baseColumn, 0)
                sheet.cell(nowLine + 2 + i * 3, j * 2 + baseColumn + 1, '失分').alignment = align
                sheet.cell(nowLine + 2 + i * 3 + 1, j * 2 + baseColumn + 1, line[shoots[j]]['lost']).alignment = align
                excelUtils.fillColorByValue(MaxScore, line[shoots[j]]['lost'], sheet, nowLine + 2 + i * 3 + 1,
                                            j * 2 + baseColumn + 1, 1)

        nowLine += 7

    nowLine += 1
    fillRed = PatternFill("solid", fgColor="f89588")
    fillRGreen = PatternFill("solid", fgColor="76da91")

    sheet.cell(nowLine, baseColumn, '发接发顺序').alignment = align
    sheet.cell(nowLine, baseColumn).font = Font(bold=True)
    sheet.merge_cells(start_row=nowLine, start_column=baseColumn, end_row=nowLine, end_column=baseColumn+3)

    order = list(calScore.keys())
    sheet.cell(nowLine, baseColumn, '发接发顺序').alignment = align
    sheet.cell(nowLine, baseColumn + 4, '得分数').alignment = align
    sheet.cell(nowLine, baseColumn + 5, '失分数').alignment = align
    sheet.cell(nowLine, baseColumn + 6, '得分率').alignment = align
    sheet.cell(nowLine, baseColumn + 7, '优劣势').alignment = align
    sheet.cell(nowLine, baseColumn).font = Font(bold=True)
    sheet.cell(nowLine, baseColumn + 4).font = Font(bold=True)
    sheet.cell(nowLine, baseColumn + 5).font = Font(bold=True)
    sheet.cell(nowLine, baseColumn + 6).font = Font(bold=True)
    sheet.cell(nowLine, baseColumn + 7).font = Font(bold=True)
    nowLine += 1
    for i in range(len(order)):
        sheet.cell(nowLine, baseColumn, order[i]).alignment = align
        sheet.merge_cells(start_row=nowLine, start_column=baseColumn, end_row=nowLine, end_column=baseColumn+3)
        sheet.cell(nowLine, baseColumn + 4, calScore[order[i]]['得分数']).alignment = align
        sheet.cell(nowLine, baseColumn + 5, calScore[order[i]]['失分数']).alignment = align
        sheet.cell(nowLine, baseColumn + 6, calScore[order[i]]['得分率']).alignment = align
        sheet.cell(nowLine, baseColumn + 7, calScore[order[i]]['优劣势']).alignment = align
        if calScore[order[i]]['优劣势'] == '劣势':
            sheet.cell(nowLine, baseColumn + 7).fill = fillRed
        elif calScore[order[i]]['优劣势'] == '优势':
            sheet.cell(nowLine, baseColumn + 7).fill = fillRGreen
        nowLine += 1
    nowLine += 1
    return nowLine
