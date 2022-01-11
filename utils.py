# 判断发球方，返回运动员下标
import os
import openpyxl
import excelUtils
import TechScoreCase


def getServeSide(point):
    if point['startSide'] != -1:
        return point['startSide']
    winSide = point['winSide']
    rallyNum = len(point['list'])
    # 偶数拍发球方win，奇数拍发球方lost
    if rallyNum % 2 == 0:
        serveSide = winSide
    else:
        serveSide = 1 if winSide == 0 else 0
    return serveSide


# 根据当前 发球落点、发球方、该point数据，推断发球方胜负，并更新 dict
def updateServeDict(serveDict, servePoint, serveSide, winSide):
    if not (servePoint in serveDict[serveSide['name']]):
        serveDict[serveSide['name']][servePoint] = {'win': 0, 'lost': 0}
    if winSide['name'] == serveSide['name']:
        serveDict[serveSide['name']][servePoint]['win'] += 1
        serveDict[serveSide['name']]['totalWin'] += 1
    else:
        serveDict[serveSide['name']][servePoint]['lost'] += 1
        serveDict[serveSide['name']]['totalLost'] += 1


#   判断 player1 的击球是斜线还是直线 diagonal | straight
#  strikePos 是 p1 挥拍的身位，point1 是 p1 接到的球的落点，point2 是 p1 打到 p2 桌上的落点
def getStrokeDir(strikePos, point1, point2, player1, player2):
    straight = 'straight'
    diagonal = 'diagonal'
    backhand = 'B'
    forehand = 'F'
    reBackhand = 'T'
    reForehand = 'P'
    if player1['rightHand'] == player2['rightHand']:  # 执拍手相同，不需要判断
        if strikePos == backhand:
            if (point1[0] == 'B' == point2[0]) or (point1[0] == 'M' and point2[0] == 'B'):
                return [backhand, diagonal]
            if (point1[0] == 'B' and point2[0] == 'F') or (point1[0] == 'B' and point2[0] == 'M') \
                    or (point1[0] == 'M' and point2[0] == 'F') or (point1[0] == point2[0] == 'M'):
                return [backhand, straight]
        elif strikePos == forehand:
            if (point1[0] == 'F' == point2[0]) or (point1[0] == 'M' and point2[0] == 'F') \
                    or (point1[0] == 'M' and point2[0] == 'B'):
                return [forehand, diagonal]
            if (point1[0] == 'F' and point2[0] == 'M') or (point1[0] == 'F' and point2[0] == 'B') \
                    or ([point1[0] == 'M' == point2[0]]):
                return [forehand, straight]
        elif strikePos == reBackhand:
            if (point1[0] == point2[0] == 'F') or (point1[0] == 'M' and point2[0] == 'B'):
                return [backhand, diagonal]
            if (point1[0] == 'F' and point2[0] == 'M') or (point1[0] == 'F' and point2[0] == 'B'):
                return [backhand, straight]
        elif strikePos == reForehand:
            if (point1[0] == 'B' == point2[0]) or (point1[0] == 'M' and point2[0] != 'M'):
                return [forehand, diagonal]
            if (point1[0] == 'B' and point2[0] == 'F') or (point1[0] == 'B' and point2[0] == 'M'):
                return [forehand, straight]
    else:  # 执拍手不同，需要判断一下
        if strikePos == backhand:
            if (point1[0] == 'B' and point2[0] == 'F') or (point1[0] == 'M' and point2[0] == 'F'):
                return [backhand, diagonal]
            if (point1[0] == 'B' and point2[0] == 'B') or (point1[0] == 'B' and point2[0] == 'M') \
                    or (point1[0] == 'M' and point2[0] == 'B') or (point1[0] == point2[0] == 'M'):
                return [backhand, straight]
        elif strikePos == forehand:
            if (point1[0] == 'F' and point2[0] == 'B') or (point1[0] == 'M' and point2[0] == 'F') \
                    or (point1[0] == 'M' and point2[0] == 'B'):
                return [forehand, diagonal]
            if (point1[0] == 'F' and point2[0] == 'M') or (point1[0] == 'F' and point2[0] == 'F') \
                    or ([point1[0] == 'M' == point2[0]]):
                return [forehand, straight]
        elif strikePos == reBackhand:
            if (point1[0] == 'F' and point2[0] == 'B') or (point1[0] == 'M' and point2[0] == 'F'):
                return [backhand, diagonal]
            if (point1[0] == 'F' and point2[0] == 'M') or (point1[0] == 'F' and point2[0] == 'F'):
                return [backhand, straight]
        elif strikePos == reForehand:
            if (point1[0] == 'B' and point2[0] == 'F') or (point1[0] == 'M' and point2[0] != 'M'):
                return [forehand, diagonal]
            if (point1[0] == 'B' and point2[0] == 'B') or (point1[0] == 'B' and point2[0] == 'M'):
                return [forehand, straight]


def updateScoreCase(scoreCaseDict, rallyList, player1, player2, serveSide, winSide):
    #   mp 前4个表示身位字符值对应的矩阵行号，后9个表示落点字符值对应的矩阵列号，整个mp用于根据字符值映射二维矩阵下标
    mp = {'B': 0, 'F': 1, 'P': 2, 'T': 3, 'BS': 0, 'BH': 1, 'BL': 2, 'MS': 3, 'MH': 4, 'ML': 5, 'FS': 6,
          'FH': 7, 'FL': 8, 'SP': 9}
    winStroke = rallyList[-2]
    lastStroke = rallyList[-1]
    winStrike = winStroke['StrikePosition']['value']
    winBallPos = lastStroke['BallPosition']['value']
    # 赢的一方的最后一拍是相持阶段的，那么 scoreCaseDict 中获胜方运动员的相持得分中，对应的（身位-落点）线路得分 +1
    if winStroke['index'] > 3:  # 双方击球大于4板认为是相持
        scoreCaseDict[winSide['name']].strike_ball_score_more[mp[winStrike]][mp[winBallPos]] += 1
    if winStroke['index'] == 3:  # 第4拍得分情况
        scoreCaseDict[winSide['name']].strike_ball_score_4[mp[winStrike]][mp[winBallPos]] += 1
    if winStroke['index'] == 2:  # 第3拍得分情况
        scoreCaseDict[winSide['name']].strike_ball_score_3[mp[winStrike]][mp[winBallPos]] += 1
    if winStroke['index'] == 1:  # 接发球得分情况
        scoreCaseDict[winSide['name']].strike_ball_score_rec[mp[winStrike]][mp[winBallPos]] += 1


def updateLineScoreCase(lineScoreCaseDict, rallyList, player1, player2, winSide, techIndex):
    winStroke = rallyList[-2]
    lastStroke = rallyList[-1]
    winStrike = winStroke['StrikePosition']['value']
    point1 = winStroke['BallPosition']['value']
    point2 = lastStroke['BallPosition']['value']
    lineDir = getStrokeDir(winStrike, point1, point2, player1, player2)
    # print(lineDir)
    mp = {1: '接发球', 2: '第三拍', 3: '第四拍', 4: '相持'}
    if lineDir is not None:
        if 3 >= techIndex > 0:
            if lineDir[0] == 'B' and lineDir[1] == 'diagonal':
                lineScoreCaseDict[winSide['name']][mp[techIndex]]['反手给斜线'] += 1
            if lineDir[0] == 'F' and lineDir[1] == 'diagonal':
                lineScoreCaseDict[winSide['name']][mp[techIndex]]['正手给斜线'] += 1
            if lineDir[0] == 'B' and lineDir[1] == 'straight':
                lineScoreCaseDict[winSide['name']][mp[techIndex]]['反手给直线'] += 1
            if lineDir[0] == 'F' and lineDir[1] == 'straight':
                lineScoreCaseDict[winSide['name']][mp[techIndex]]['正手给直线'] += 1
            lineScoreCaseDict[winSide['name']][mp[techIndex]]['total'] += 1
        elif techIndex > 3:
            if lineDir[0] == 'B' and lineDir[1] == 'diagonal':
                lineScoreCaseDict[winSide['name']]['相持']['反手给斜线'] += 1
            if lineDir[0] == 'F' and lineDir[1] == 'diagonal':
                lineScoreCaseDict[winSide['name']]['相持']['正手给斜线'] += 1
            if lineDir[0] == 'B' and lineDir[1] == 'straight':
                lineScoreCaseDict[winSide['name']]['相持']['反手给直线'] += 1
            if lineDir[0] == 'F' and lineDir[1] == 'straight':
                lineScoreCaseDict[winSide['name']]['相持']['正手给直线'] += 1
            lineScoreCaseDict[winSide['name']]['相持']['total'] += 1


def createExcel(fileName, serveDict, scoreCaseDict, lineScoreCaseDict,
                playerNameList):  # playerNameList, serveDict, scoreCaseDict, lineScoreCaseDict
    #   去除文件后缀名
    (filePath, tmpFileName) = os.path.split(fileName)
    (fileName, fileType) = os.path.splitext(tmpFileName)
    typeDict = {'BS': '反手短', 'BH': '反手半出台', 'BL': '反手长', 'MS': '中路短', 'MH': '中路半出台',
                'ML': '中路长', 'FS': '正手短', 'FH': '正手半出台', 'FL': '正手长'}

    for playerName in serveDict.keys():
        exl = openpyxl.Workbook()
        sheet = exl.create_sheet('Sheet1', 0)  # 打开该Excel里对应的sheet
        #   ---------- 填入发球部分 ----------
        posTypeList = list(serveDict[playerName].keys())[2:]  # 发球出现的所有落点种类
        posTypeList.sort()
        # print(posTypeList)
        sheet.cell(2, 1, '发球落点')
        sheet.cell(2, 2, '得分')
        sheet.cell(3, 2, '失分')
        for i in range(len(posTypeList) + 1):  # 枚举该运动员发球出现的所有落点种类，写入发球表头
            if i == len(posTypeList):
                sheet.cell(1, i + 3, '合计')
                sheet.cell(2, i + 3, serveDict[playerName]['totalWin'])
                sheet.cell(3, i + 3, serveDict[playerName]['totalLost'])
                break
            sheet.cell(1, i + 3, typeDict[posTypeList[i]])
            sheet.cell(2, i + 3, serveDict[playerName][posTypeList[i]]['win'])
            sheet.cell(3, i + 3, serveDict[playerName][posTypeList[i]]['lost'])

        #   ---------- 填入线路左侧部分 ----------
        beginLine = 5
        cur = 0
        for techIndex in list(lineScoreCaseDict[playerName].keys()):  # techIndex 为接发球、第三拍、相持等等
            sheet.cell(beginLine + cur * 3 + 1, 1, techIndex if techIndex != '相持' else playerName + techIndex)
            sheet.cell(beginLine + cur * 3 + 1, 2, '得分')
            lineTypeList = list(lineScoreCaseDict[playerName][techIndex].keys())  # lineTypeList 元素为线路种类，如反手给斜线
            for i in range(len(lineTypeList)):
                sheet.cell(beginLine + cur * 3, i + 3, lineTypeList[i] if lineTypeList[i] != 'total' else '合计')
            for i in range(len(lineTypeList)):  # 枚举该运动员发球出现的所有落点种类，写入发球表头
                sheet.cell(beginLine + cur * 3 + 1, i + 3, lineScoreCaseDict[playerName][techIndex][lineTypeList[i]])
            cur += 1
        # 相持的统计需要在下一行加上对方选手的
        anotherPlayerName = ''
        for name in playerNameList:
            if name != playerName:
                anotherPlayerName = name
                break
        sheet.cell(beginLine + cur * 3 - 1, 1, anotherPlayerName + '相持')
        sheet.cell(beginLine + cur * 3 - 1, 2, '得分')
        lineTypeList = list(lineScoreCaseDict[anotherPlayerName]['相持'].keys())  # lineTypeList 元素为线路种类，如反手给斜线
        for i in range(len(lineTypeList)):  # 枚举该运动员发球出现的所有落点种类，写入发球表头
            sheet.cell(beginLine + cur * 3 - 1, i + 3, lineScoreCaseDict[anotherPlayerName]['相持'][lineTypeList[i]])

        #   ---------- 填入线路左侧部分 ----------
        strikeMp = {0: '反手', 1: '正手', 2: '侧身', 3: '反侧身'}
        ballMp = {0: '反手短', 1: '反手半出台', 2: '反手长', 3: '中路短', 4: '中路半出台',
                  5: '中路长', 6: '正手短', 7: '正手半出台', 8: '正手长', 9: '擦网擦边'}
        tmp = scoreCaseDict[playerName]
        techIndexList = [tmp.strike_ball_score_rec, tmp.strike_ball_score_3, tmp.strike_ball_score_4,
                         tmp.strike_ball_score_more]
        beginLine = 5
        cur = 0
        for techIndex in techIndexList:
            for i in range(4):
                for j in range(10):
                    sheet.cell(beginLine + cur * 3, i * 10 + j + len(lineTypeList) + 4, strikeMp[i] + '给' + ballMp[j])
                    sheet.cell(beginLine + cur * 3 + 1, i * 10 + j + len(lineTypeList) + 4, techIndex[i][j])
                sheet.cell(beginLine + cur * 3, i * 10 + 10 + len(lineTypeList) + 4, '合计')
                sheet.cell(beginLine + cur * 3 + 1, i * 10 + 10 + len(lineTypeList) + 4, TechScoreCase.calSum(techIndex))
            cur += 1

        excelUtils.style_excel(sheet)

        exl.save(fileName + '(' + playerName + '习惯性出手线路情况)' + '.xlsx')


# gameProgress = [
#     0, 0, 0, 0, 0, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,
#     0, 0, 0, 0, 0, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,
#     0, 0, 0, 0, 0, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,
#     0, 0, 0, 0, 0, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,
#     0, 0, 0, 0, 0, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,
#     1, 1, 1, 1, 1, 1, 1, 1, 1, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,
#     1, 1, 1, 1, 1, 1, 1, 1, 1, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,
#     1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 1, 1, 1, 1, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 1, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
#     2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
gameProgress = [
    [0, 0, 0, 0, 0, 1, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [0, 0, 0, 0, 0, 1, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [0, 0, 0, 0, 0, 1, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [0, 0, 0, 0, 0, 1, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [0, 0, 0, 0, 0, 1, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [1, 1, 1, 1, 1, 1, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [1, 1, 1, 1, 1, 1, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [1, 1, 1, 1, 1, 1, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [1, 1, 1, 1, 1, 1, 1, 1, 1, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],

    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2],
    [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,     2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2]
]
# ]


#   0：常规僵持    1：落后    2：领先    3：失控    4：掌控  5：关键球僵持
gameState = [
    [0, 0, 0, 1, 1, 3, 3, 3, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [0, 0, 0, 0, 1, 1, 3, 3, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [0, 0, 0, 0, 0, 1, 1, 3, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [2, 0, 0, 0, 0, 0, 1, 1, 3, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [2, 2, 0, 0, 0, 0, 0, 1, 1, 3, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [4, 2, 2, 0, 0, 0, 0, 0, 1, 1, 3, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [4, 4, 2, 2, 0, 0, 0, 0, 0, 1, 1, 3,     3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [4, 4, 4, 2, 2, 0, 0, 0, 0, 0, 1, 1,     1, 1, 1, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [4, 4, 4, 4, 2, 2, 0, 0, 0, 5, 5, 1,     1, 1, 1, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [4, 4, 4, 4, 4, 2, 2, 0, 5, 5, 5, 5,     1, 1, 1, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [4, 4, 4, 4, 4, 4, 2, 2, 5, 5, 5, 5,     5, 1, 1, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],
    [4, 4, 4, 4, 4, 4, 4, 2, 2, 5, 5, 5,     5, 5, 1, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3],


    [4, 4, 4, 4, 4, 4, 4, 4, 2, 2, 5, 5,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 2, 2, 5,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 2, 2,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 2,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
    [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4,     5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5],
]


