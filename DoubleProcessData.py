import openpyxl
import json
import DoubleUtils
import utils

gameProgress = utils.gameProgress
gameState = utils.gameState
progressNames = ['开局', '中局', '尾局']
stateNames = ['常规僵持', '落后', '领先', '失控', '掌控', '关键球僵持']

def notProgress(fileNameList):
    nowLine = 1
    fileIdx = 0
    exl = openpyxl.Workbook()
    sheet = exl.create_sheet('Sheet1', 0)  # 打开该Excel里对应的sheet

    # fileNameList = ['20210726 东京奥运会 混双决赛 许昕刘诗雯vs水谷隼伊藤美诚-collect_project.json']
    for fileName in fileNameList:
        serve_rec_order = {'00': '', '01': '', '10': '', '11': ''}
        fileIdx += 1

        print(fileName)
        with open(fileName, 'r', encoding='utf-8') as f:
            jsonFile = f.read()
            jsonFile = json.loads(jsonFile)

        data = jsonFile['data']
        playerList = data['player']
        team1 = playerList[0]
        team2 = playerList[1]
        print(playerList, team1, team2)

        #   得到统计对象的TeamID（str类型）
        TargetTeamPair = DoubleUtils.getTargetTeamID(playerList)

        roundList = data['record']['list']  # 若干局，元素是每个个大比分 局内的的信息

        hitOrders = DoubleUtils.getHitOrders(playerList)
        print(hitOrders)
        # scoreCase在双打时有8个元素，每个元素是一个[]， 里面有2个dict，分别是2种开球情况，dict里面第x个元素存放第x拍的得失分{'win': 0, 'lose': 0}
        scoreCase = [[{}, {}] for i in range(len(hitOrders))]
        scoreCaseOpposite = [[{}, {}] for i in range(len(hitOrders))]
        print(scoreCase)
        matchResult = data['record']['result']
        isFull = (abs(matchResult[0] - matchResult[1]) == 1)  # 是否打到决胜局
        print('res', matchResult, isFull)

        count = 0
        # print(DoubleUtils.getHitOrders(playerList))

        for roundId, round in enumerate(roundList):  # 其中的一局的信息
            cnt = 0
            cnt2 = 0

            pointList = round['list']  # pointList 中元素是该局的每一分的信息
            # print(playerList[utils.getServeSide(pointList[0])][0]['name'])  # 测试判断发球方的函数
            for pointId, point in enumerate(pointList):
                count += 1
                rallyList = point['list']  # rallyList 记录该分中所有挥拍，每个元素是一个挥拍
                rallyNum = len(rallyList)

                #   单打情况--------------------------------------------------
                if len(hitOrders) == 2:
                    playerPair = ['00', '10']
                    A1 = playerPair[DoubleUtils.getTargetPlayer(playerPair, TargetTeamPair[0])]  # 当前pair中中国运动员
                    B1 = playerPair[DoubleUtils.getOppositePlayer(playerPair, TargetTeamPair[0])]  # 当前pair对方运动员

                    if rallyNum == 1:  # 发球失误的情况特判一下
                        if rallyList[0]['HitPlayer'] == A1:
                            print('发球失误', A1, B1)
                            index = hitOrders.index([A1, B1])
                            DoubleUtils.updateScoreCase(scoreCase, index, 0, 0, 'lost')
                            DoubleUtils.updateScoreCase(scoreCase, index, 1, 1, 'win')
                        elif rallyList[0]['HitPlayer'] == B1:
                            print('发球失误', B1, A1)
                            index = hitOrders.index([B1, A1])
                            DoubleUtils.updateScoreCase(scoreCase, index, 1, 1, 'win')
                            DoubleUtils.updateScoreCase(scoreCase, index, 0, 0, 'lost')
                        break
                    servePair = [rallyList[0]['HitPlayer'], rallyList[1]['HitPlayer']]
                    lastHitPair = [rallyList[-2]['HitPlayer'], rallyList[-1]['HitPlayer']]
                    index = hitOrders.index(servePair)
                    another = 1 if index == 0 else 0
                    if servePair[0] == lastHitPair[0]:  # 发球的赢了
                        DoubleUtils.updateScoreCase(scoreCase, index, 0, rallyNum - 2, 'win')
                        DoubleUtils.updateScoreCase(scoreCase, another, 1, rallyNum - 1, 'lost')
                    else:
                        DoubleUtils.updateScoreCase(scoreCase, index, 0, rallyNum - 1, 'lost')
                        DoubleUtils.updateScoreCase(scoreCase, another, 1, rallyNum - 2, 'win')
                    continue

                #   双打情况--------------------------------------------------
                if roundId == 0 and pointId == 0:
                    DoubleUtils.updateServeRecOrder(serve_rec_order, rallyList[0]['HitPlayer'],
                                                    rallyList[1]['HitPlayer'])
                if roundId != 0 and pointId == 0:
                    DoubleUtils.updateServeRecOrder(serve_rec_order, rallyList[0]['HitPlayer'])

                if roundId == len(roundList) - 1 and isFull:  # 决胜局到5分换次序情况
                    if max(point['score']) == 5:
                        DoubleUtils.updateServeRecOrder(serve_rec_order, rallyList[0]['HitPlayer'])

                DoubleUtils.updateDoublesData(scoreCase, rallyList, hitOrders, TargetTeamPair, serve_rec_order, False,
                                              cnt, cnt2)
                DoubleUtils.updateDoublesData(scoreCaseOpposite, rallyList, hitOrders, TargetTeamPair, serve_rec_order,
                                              True,
                                              cnt, cnt2)

            print(cnt, cnt2)

        print(scoreCase)
        print(scoreCaseOpposite)
        print('total point = ', count)

        excelScoreCase = {}
        excelScoreCaseOpposite = {}
        calScore = {}
        calScoreOpposite = {}
        count = 0
        MaxScore = DoubleUtils.getExcelScoreCase(excelScoreCase, scoreCase, calScore, hitOrders, playerList, False)
        MaxScoreOpposite = DoubleUtils.getExcelScoreCase(excelScoreCaseOpposite, scoreCaseOpposite,
                                                         calScoreOpposite, hitOrders, playerList, True)

        # ----------------- 每局下方不同轮次的得分情况、得分率统计 ------------------
        DoubleUtils.getAdvanceRate(excelScoreCase, scoreCase, calScore, hitOrders, playerList)
        DoubleUtils.getAdvanceRate(excelScoreCaseOpposite, scoreCaseOpposite, calScoreOpposite, hitOrders, playerList)
        # print(count)

        baseColumn = 1
        baseColumnOpposite = 16

        tmpLine = nowLine
        nowLine = DoubleUtils.fillExcel(excelScoreCase, calScore, tmpLine, baseColumn, sheet, fileName, MaxScore)
        if len(hitOrders) != 2:  # 单打情况不需要另一半，另一半在 hitOders 的另一种就涵盖了
            nowLine = DoubleUtils.fillExcel(excelScoreCaseOpposite, calScoreOpposite, tmpLine,
                                            baseColumnOpposite, sheet, fileName, MaxScoreOpposite)

    exl.save('双打得失分数据统计（共' + str(fileIdx) + '场比赛）' + '.xlsx')


def byProgress(fileNameList):
    exl = openpyxl.Workbook()
    for progress in range(3):
        nowLine = 1
        fileIdx = 0
        sheet = exl.create_sheet(progressNames[progress], progress)  # 打开该Excel里对应的sheet

        # fileNameList = ['20210726 东京奥运会 混双决赛 许昕刘诗雯vs水谷隼伊藤美诚-collect_project.json']
        for fileName in fileNameList:
            serve_rec_order = {'00': '', '01': '', '10': '', '11': ''}
            fileIdx += 1

            print(fileName)
            with open(fileName, 'r', encoding='utf-8') as f:
                jsonFile = f.read()
                jsonFile = json.loads(jsonFile)

            data = jsonFile['data']
            playerList = data['player']
            team1 = playerList[0]
            team2 = playerList[1]
            print(playerList, team1, team2)

            #   得到统计对象的TeamID（str类型）
            TargetTeamPair = DoubleUtils.getTargetTeamID(playerList)

            roundList = data['record']['list']  # 若干局，元素是每个个大比分 局内的的信息

            hitOrders = DoubleUtils.getHitOrders(playerList)
            # print(hitOrders)

            # scoreCase在双打时有8个元素，每个元素是一个[]， 里面有2个dict，分别是2种开球情况，dict里面第x个元素存放第x拍的得失分{'win': 0, 'lose': 0}
            scoreCase = [[{}, {}] for i in range(len(hitOrders))]
            scoreCaseOpposite = [[{}, {}] for i in range(len(hitOrders))]
            # print(scoreCase)
            matchResult = data['record']['result']
            isFull = (abs(matchResult[0] - matchResult[1]) == 1)  # 是否打到决胜局
            print('res', matchResult, isFull)

            count = 0
            # print(DoubleUtils.getHitOrders(playerList))

            for roundId, round in enumerate(roundList):  # 其中的一局的信息
                cnt = 0
                cnt2 = 0

                pointList = round['list']  # pointList 中元素是该局的每一分的信息
                # print(playerList[utils.getServeSide(pointList[0])][0]['name'])  # 测试判断发球方的函数
                for pointId, point in enumerate(pointList):
                    count += 1
                    rallyList = point['list']  # rallyList 记录该分中所有挥拍，每个元素是一个挥拍
                    rallyNum = len(rallyList)

                    #   单打情况--------------------------------------------------
                    if len(hitOrders) == 2:
                        playerPair = ['00', '10']
                        A1 = playerPair[DoubleUtils.getTargetPlayer(playerPair, TargetTeamPair[0])]  # 当前pair中中国运动员
                        B1 = playerPair[DoubleUtils.getOppositePlayer(playerPair, TargetTeamPair[0])]  # 当前pair对方运动员

                        if rallyNum == 1:  # 发球失误的情况特判一下
                            if rallyList[0]['HitPlayer'] == A1:
                                print('发球失误', A1, B1)
                                index = hitOrders.index([A1, B1])
                                DoubleUtils.updateScoreCase(scoreCase, index, 0, 0, 'lost')
                                DoubleUtils.updateScoreCase(scoreCase, index, 1, 1, 'win')
                            elif rallyList[0]['HitPlayer'] == B1:
                                print('发球失误', B1, A1)
                                index = hitOrders.index([B1, A1])
                                DoubleUtils.updateScoreCase(scoreCase, index, 1, 1, 'win')
                                DoubleUtils.updateScoreCase(scoreCase, index, 0, 0, 'lost')
                            break
                        servePair = [rallyList[0]['HitPlayer'], rallyList[1]['HitPlayer']]
                        lastHitPair = [rallyList[-2]['HitPlayer'], rallyList[-1]['HitPlayer']]
                        index = hitOrders.index(servePair)
                        another = 1 if index == 0 else 0
                        if servePair[0] == lastHitPair[0]:  # 发球的赢了
                            DoubleUtils.updateScoreCase(scoreCase, index, 0, rallyNum - 2, 'win')
                            DoubleUtils.updateScoreCase(scoreCase, another, 1, rallyNum - 1, 'lost')
                        else:
                            DoubleUtils.updateScoreCase(scoreCase, index, 0, rallyNum - 1, 'lost')
                            DoubleUtils.updateScoreCase(scoreCase, another, 1, rallyNum - 2, 'win')
                        continue

                    #   双打情况--------------------------------------------------
                    if roundId == 0 and pointId == 0:
                        DoubleUtils.updateServeRecOrder(serve_rec_order, rallyList[0]['HitPlayer'],
                                                        rallyList[1]['HitPlayer'])
                    if roundId != 0 and pointId == 0:
                        DoubleUtils.updateServeRecOrder(serve_rec_order, rallyList[0]['HitPlayer'])

                    if roundId == len(roundList) - 1 and isFull:  # 决胜局到5分换次序情况
                        if max(point['score']) == 5:
                            DoubleUtils.updateServeRecOrder(serve_rec_order, rallyList[0]['HitPlayer'])

# ---------------------------------------------------------------------------------------------------------------
                    nowScore = [point['score'][0], point['score'][1]]
                    tmp = [nowScore[0], nowScore[1]]
                    nowScore[point['winSide']] += 1
                    if gameProgress[nowScore[0]][nowScore[1]] != gameProgress[tmp[0]][tmp[1]]:
                        print('临界得分：', '局数：', roundId + 1, tmp, nowScore)
                    if gameProgress[nowScore[0]][nowScore[1]] != progress:
                        continue
                    print('当前局分', nowScore, 'id', progress, '态势', progressNames[progress])

                    # ---------------------------------------------------------------------------------------------------------------

                    DoubleUtils.updateDoublesData(scoreCase, rallyList, hitOrders, TargetTeamPair, serve_rec_order,
                                                  False,
                                                  cnt, cnt2)
                    DoubleUtils.updateDoublesData(scoreCaseOpposite, rallyList, hitOrders, TargetTeamPair,
                                                  serve_rec_order,
                                                  True,
                                                  cnt, cnt2)

                print(cnt, cnt2)

            print(scoreCase)
            print(scoreCaseOpposite)
            print('total point = ', count)

            excelScoreCase = {}
            excelScoreCaseOpposite = {}
            calScore = {}
            calScoreOpposite = {}
            count = 0
            MaxScore = DoubleUtils.getExcelScoreCase(excelScoreCase, scoreCase, calScore, hitOrders, playerList, False)
            MaxScoreOpposite = DoubleUtils.getExcelScoreCase(excelScoreCaseOpposite, scoreCaseOpposite,
                                                             calScoreOpposite, hitOrders, playerList, True)

            # ----------------- 每局下方不同轮次的得分情况、得分率统计 ------------------
            DoubleUtils.getAdvanceRate(excelScoreCase, scoreCase, calScore, hitOrders, playerList)
            DoubleUtils.getAdvanceRate(excelScoreCaseOpposite, scoreCaseOpposite, calScoreOpposite, hitOrders,
                                       playerList)
            # print(count)

            baseColumn = 1
            baseColumnOpposite = 16

            tmpLine = nowLine
            nowLine = DoubleUtils.fillExcel(excelScoreCase, calScore, tmpLine, baseColumn, sheet, fileName, MaxScore)
            if len(hitOrders) != 2:  # 单打情况不需要另一半，另一半在 hitOders 的另一种就涵盖了
                nowLine = DoubleUtils.fillExcel(excelScoreCaseOpposite, calScoreOpposite, tmpLine,
                                                baseColumnOpposite, sheet, fileName, MaxScoreOpposite)

    exl.save('（分阶段）双打得失分数据统计（共' + str(fileIdx) + '场比赛）' + '.xlsx')


def byState(fileNameList):
    exl = openpyxl.Workbook()
    fileIdx = 0
    # fileNameList = ['20210726 东京奥运会 混双决赛 许昕刘诗雯vs水谷隼伊藤美诚-collect_project.json']
    for fileName in fileNameList:
        serve_rec_order = {'00': '', '01': '', '10': '', '11': ''}
        fileIdx += 1
        print(fileName)
        with open(fileName, 'r', encoding='utf-8') as f:
            jsonFile = f.read()
            jsonFile = json.loads(jsonFile)

        data = jsonFile['data']
        playerList = data['player']
        team1 = playerList[0]
        team2 = playerList[1]

        #   得到统计对象的TeamID（str类型）
        TargetTeamPair = DoubleUtils.getTargetTeamID(playerList)

        roundList = data['record']['list']  # 若干局，元素是每个个大比分 局内的的信息

        hitOrders = DoubleUtils.getHitOrders(playerList)
        print(hitOrders)
        matchResult = data['record']['result']
        isFull = (abs(matchResult[0] - matchResult[1]) == 1)  # 是否打到决胜局
        print('res', matchResult, isFull)

        count = 0

        for state in range(len(stateNames)):
            nowLine = 1
            sheet = exl.create_sheet(stateNames[state], state)  # 打开该Excel里对应的sheet
            # scoreCase在双打时有8个元素，每个元素是一个[]， 里面有2个dict，分别是2种开球情况，dict里面第x个元素存放第x拍的得失分{'win': 0, 'lose': 0}
            scoreCase = [[{}, {}] for i in range(len(hitOrders))]
            scoreCaseOpposite = [[{}, {}] for i in range(len(hitOrders))]
            print(scoreCase)
            for roundId, round in enumerate(roundList):  # 其中的一局的信息
                cnt = 0
                cnt2 = 0

                pointList = round['list']  # pointList 中元素是该局的每一分的信息
                # print(playerList[utils.getServeSide(pointList[0])][0]['name'])  # 测试判断发球方的函数
                for pointId, point in enumerate(pointList):
                    count += 1
                    rallyList = point['list']  # rallyList 记录该分中所有挥拍，每个元素是一个挥拍
                    rallyNum = len(rallyList)

                    #   单打情况--------------------------------------------------
                    if len(hitOrders) == 2:
                        playerPair = ['00', '10']
                        A1 = playerPair[DoubleUtils.getTargetPlayer(playerPair, TargetTeamPair[0])]  # 当前pair中中国运动员
                        B1 = playerPair[DoubleUtils.getOppositePlayer(playerPair, TargetTeamPair[0])]  # 当前pair对方运动员

                        if rallyNum == 1:  # 发球失误的情况特判一下
                            if rallyList[0]['HitPlayer'] == A1:
                                print('发球失误', A1, B1)
                                index = hitOrders.index([A1, B1])
                                DoubleUtils.updateScoreCase(scoreCase, index, 0, 0, 'lost')
                                DoubleUtils.updateScoreCase(scoreCase, index, 1, 1, 'win')
                            elif rallyList[0]['HitPlayer'] == B1:
                                print('发球失误', B1, A1)
                                index = hitOrders.index([B1, A1])
                                DoubleUtils.updateScoreCase(scoreCase, index, 1, 1, 'win')
                                DoubleUtils.updateScoreCase(scoreCase, index, 0, 0, 'lost')
                            break
                        servePair = [rallyList[0]['HitPlayer'], rallyList[1]['HitPlayer']]
                        lastHitPair = [rallyList[-2]['HitPlayer'], rallyList[-1]['HitPlayer']]
                        index = hitOrders.index(servePair)
                        another = 1 if index == 0 else 0
                        if servePair[0] == lastHitPair[0]:  # 发球的赢了
                            DoubleUtils.updateScoreCase(scoreCase, index, 0, rallyNum - 2, 'win')
                            DoubleUtils.updateScoreCase(scoreCase, another, 1, rallyNum - 1, 'lost')
                        else:
                            DoubleUtils.updateScoreCase(scoreCase, index, 0, rallyNum - 1, 'lost')
                            DoubleUtils.updateScoreCase(scoreCase, another, 1, rallyNum - 2, 'win')
                        continue

                    #   双打情况--------------------------------------------------
                    if roundId == 0 and pointId == 0:
                        DoubleUtils.updateServeRecOrder(serve_rec_order, rallyList[0]['HitPlayer'],
                                                        rallyList[1]['HitPlayer'])
                    if roundId != 0 and pointId == 0:
                        DoubleUtils.updateServeRecOrder(serve_rec_order, rallyList[0]['HitPlayer'])

                    if roundId == len(roundList) - 1 and isFull:  # 决胜局到5分换次序情况
                        if max(point['score']) == 5:
                            DoubleUtils.updateServeRecOrder(serve_rec_order, rallyList[0]['HitPlayer'])

                    # ---------------------------------------------------------------------------------------------------------------
                    nowScore = [point['score'][0], point['score'][1]]
                    print(nowScore)
                    tmp = [nowScore[0], nowScore[1]]
                    nowScore[point['winSide']] += 1
                    if gameProgress[nowScore[0]][nowScore[1]] != gameProgress[tmp[0]][tmp[1]]:
                        print('临界得分：', '局数：', roundId + 1, tmp, nowScore)
                    # print(nowScore)
                    teamIdPair = list(map(int, DoubleUtils.getTargetTeamID(playerList)))
                    # if gameState[nowScore[teamIdPair[0]]][nowScore[teamIdPair[1]]] != state:
                        # continue
                    # ---------------------------------------------------------------------------------------------------------------
                    if gameState[nowScore[teamIdPair[0]]][nowScore[teamIdPair[1]]] == state:
                        print('当前局分', nowScore, 'id', state, '态势', stateNames[state])
                        DoubleUtils.updateDoublesData(scoreCase, rallyList, hitOrders, TargetTeamPair, serve_rec_order,
                                                  False,
                                                  cnt, cnt2)
                    if gameState[nowScore[teamIdPair[1]]][nowScore[teamIdPair[0]]] == state:
                        print('当前局分', nowScore, 'id', state, '态势', stateNames[state])
                        DoubleUtils.updateDoublesData(scoreCaseOpposite, rallyList, hitOrders, TargetTeamPair,
                                                  serve_rec_order,
                                                  True,
                                                  cnt, cnt2)

                print(cnt, cnt2)

            print(scoreCase)
            print(scoreCaseOpposite)
            print('total point = ', count)

            excelScoreCase = {}
            excelScoreCaseOpposite = {}
            calScore = {}
            calScoreOpposite = {}
            count = 0
            MaxScore = DoubleUtils.getExcelScoreCase(excelScoreCase, scoreCase, calScore, hitOrders, playerList, False)
            MaxScoreOpposite = DoubleUtils.getExcelScoreCase(excelScoreCaseOpposite, scoreCaseOpposite,
                                                             calScoreOpposite, hitOrders, playerList, True)

            # ----------------- 每局下方不同轮次的得分情况、得分率统计 ------------------
            DoubleUtils.getAdvanceRate(excelScoreCase, scoreCase, calScore, hitOrders, playerList)
            DoubleUtils.getAdvanceRate(excelScoreCaseOpposite, scoreCaseOpposite, calScoreOpposite, hitOrders,
                                       playerList)
            # print(count)

            baseColumn = 1
            baseColumnOpposite = 16

            tmpLine = nowLine
            nowLine = DoubleUtils.fillExcel(excelScoreCase, calScore, tmpLine, baseColumn, sheet, fileName, MaxScore)
            if len(hitOrders) != 2:  # 单打情况不需要另一半，另一半在 hitOders 的另一种就涵盖了
                nowLine = DoubleUtils.fillExcel(excelScoreCaseOpposite, calScoreOpposite, tmpLine,
                                                baseColumnOpposite, sheet, fileName, MaxScoreOpposite)

    exl.save('（分态势）双打得失分数据统计（共' + str(fileIdx) + '场比赛）' + '.xlsx')


def process(fileNameList, callback, isProgress):
    print('isProgress', isProgress)
    if isProgress == 0:
        notProgress(fileNameList)
    elif isProgress == 1:
        byProgress(fileNameList)
    elif isProgress == 2:
        byState(fileNameList)


